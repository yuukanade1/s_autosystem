# from mysqlx import Row
import pyautogui as pg
import time
import openpyxl
import datetime
import subprocess
import win32gui
import win32con
import win32api
import pyperclip
from tkinter import messagebox

subprocess.Popen(r'C:\Windows\notepad.exe')
time.sleep(2)
def foreground():
    # hwnd = win32gui.FindWindow(None, '受注情報登録（ベルト）')
    hwnd = win32gui.FindWindow(None, '無題 - メモ帳')
    win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
    left, top, right, bottom = win32gui.GetWindowRect(hwnd)
    pg.moveTo(left + 60, top + 10)
    pg.click()
    win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)

if __name__ == '__main__':
    foreground()

dt_now = datetime.datetime.now()
print(dt_now.strftime('%Y%m%d'))

wb = openpyxl.load_workbook(r'C:\Users\竜馬\OneDrive\share\集計転記\syuukeiin.xlsx')
# wb = openpyxl.load_workbook(r'C:\Users\stock\OneDrive\share\集計転記\syuukeiin.xlsx')
ws = wb.worksheets[0]
maxClm = ws.max_column
maxRow = ws.max_row
print('maxcolumn : ' + str(maxClm))
print('maxrow    : ' + str(maxRow))
for j in range(2, maxRow + 1):
    for i in reversed(range(1,maxClm + 1)):
        if ws.cell(j, i).value != None:
            for clm in ws.iter_rows(min_row=2, max_row=j, min_col=1, max_col=i):
                values = []
                for row in clm:
                    values.append(row.value)
            time.sleep(1)
            pg.write(str(values[0]))
            pg.press('enter')
            pg.press('0')
            pg.press('enter')
            pg.write(dt_now.strftime('%Y%m%d'))
            pg.press('enter')
            pyperclip.copy(values[1])
            pg.hotkey('ctrl', 'v')
            pg.press('enter')
            for k in range(2, i):
                pg.write(str(values[k]))
                pg.press('enter')
            # pg.press('F7')
            pg.write('F7')
            pg.press('enter')
            time.sleep(1)
            print(values)
            break

wb.close()
messagebox.showinfo('自動入力', '処理が完了しました。')