from tkinter.filedialog import Open
import pyautogui as pg
import time
import openpyxl
import datetime
import subprocess
import win32gui
import win32con
import win32api
import pyperclip
from tkinter import messagebox

# column C, G, K, O, S
# clm num 3, 7, 11, 15, 19
# row 3, 13, 14, 15, 16, 17, 18, 19

# subprocess.Popen(r'C:\Windows\notepad.exe')
# time.sleep(2)
# def foreground():
#     hwnd = win32gui.FindWindow(None, '無題 - メモ帳')
#     win32gui.SetWindowPos(hwnd, win32con.HWND_TOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)
#     left, top, right, bottom = win32gui.GetWindowRect(hwnd)
#     pg.moveTo(left + 60, top + 10)
#     pg.click()
#     win32gui.SetWindowPos(hwnd, win32con.HWND_NOTOPMOST, 0, 0, 0, 0, win32con.SWP_NOMOVE | win32con.SWP_NOSIZE)

# if __name__ == '__main__':
#     foreground()

wb1 = openpyxl.load_workbook(r'C:\Users\竜馬\OneDrive\share\集計転記\自動集計s\仮2022年12月上旬_入在有_パート1.xlsx')
wb1ws = wb1.worksheets[0]

wb2 = openpyxl.load_workbook(r'C:\Users\竜馬\OneDrive\share\集計転記\syuukeiin.xlsx')
wb2ws = wb2.worksheets[0]
maxClm = wb2ws.max_column
maxRow = wb2ws.max_row
print('maxcolumn : ' + str(maxClm))
print('maxrow    : ' + str(maxRow))

clm = [3, 5, 7, 9, 11, 13, 15, 17, 19, 21]
row = [3, 13, 14, 15, 16, 17, 18, 19]

print('使用する列 : ' + str(clm))
print('使用する行 : ' + str(row))


values = []
if wb1ws.cell(row[0], clm[0]).value == None:
    values.append(wb1ws.cell(row[0], clm[0] + 1).value)
else:
    values.append(wb1ws.cell(row[0], clm[0]).value)
for i in range(1, 7, 1):
    if wb1ws.cell(row[i], clm[1]).value != None:
        values.append(wb1ws.cell(row[i], clm[0]).value)
        values.append(wb1ws.cell(row[i], clm[1]).value)

for j in range(0, len(values) - 1, 1):
    wb2ws.cell(maxRow + 1, j + 4).value = values[j]
print(values)


wb1.close()
wb2.close()
# messagebox.showinfo('自動入力', '処理が完了しました。')